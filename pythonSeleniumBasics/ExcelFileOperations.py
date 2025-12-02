import openpyxl

def readUsernameData():
    book = openpyxl.load_workbook("C:\\Users\\Raffizz\\Documents\\Obsqura\\Automation\\FileOperations.xlsx")
    sheet = book.active
    if "loginpage" in book.sheetnames:
        sheet = book.worksheets[0]
        username = sheet.cell(row=2, column=2).value
        #print(username)
    return username

print(readUsernameData())